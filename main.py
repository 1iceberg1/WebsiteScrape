import os
from openpyxl.utils import exceptions
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time
import sched
from Analyzer import Analyzer
from Record import Record
import re
import datetime
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import shutil

class MatchResult():
    def __init__(self):
        self.team1 = ""
        self.team2 = ""
        self.score1 = 0
        self.score2 = 0

UserName = 'nextaa'
PassWord = 'Qwer4321'

def check_file_writable(file_path):
    if not os.path.exists(file_path):
        return False
    try:
        # Check if the file is open
        with open(file_path, 'r') as file:
            pass  # File is not open or read-only
    except PermissionError:
        return False  # File is open or read-only
    except FileNotFoundError:
        return False  # File does not exist
    return True  # File is not open or read-only

def change_date_format(date_str, day_diff, input_format, output_format):
    date_obj = datetime.datetime.strptime(date_str, input_format)
    date_obj = date_obj + datetime.timedelta(days = day_diff)
    output_date = date_obj.strftime(output_format)
    return output_date

def get_date_value(page):
    soup = BeautifulSoup(page, 'html.parser')

    date_value = str(soup.find('span', {'id' : 'timecontainer'}).text)
    date_value = "".join(date_value.split())

    date_value = date_value[:-5]
    if len(date_value) == 18:
        if date_value[1].isdigit():
            if date_value[9] == '0':
                date_value = date_value[:9] + "12" + date_value[10:]
            else:
                date_value = date_value[:9] + "0" + date_value[9:]
        else:
            date_value = "0" + date_value
    elif len(date_value) == 17:
        if date_value[8] =='0':
            date_value = "0" + date_value[:8] + "12" + date_value[9:]
        else:
            date_value = "0" + date_value[:8] + "0" + date_value[8:]

    input_format = "%d%b%Y%I:%M:%S%p"
    output_format = "%m/%d/%Y, %I:%M:%S%p"

    date_obj = datetime.datetime.strptime(date_value, input_format)
    formatted_date_str = date_obj.strftime(output_format)
    
    return formatted_date_str

def save_workbook(filename, workbook):
    success_flag = True
    try:
        time_a = time.time()
        workbook.save(filename)
        time_b = time.time()
        print("Saving Excel costs " + str(time_b - time_a) + " seconds")
    except PermissionError:
        print("Failed to save the workbook. The file is currently open.")
        success_flag = False
    except exceptions.ReadOnlyWorkbookException:
        print("Failed to save the workbook. The file is opened as read-only.")
        success_flag = False
    except Exception as e:
        print("An error occurred while saving the workbook:", str(e))
        success_flag = False

    if success_flag == False:
        print("Please check your file is currently open or opend as read-only")

def calculate_passed_minutes(time_str):
    # Split the time string into hours, minutes, seconds, and AM/PM components
    print("TIME: " + time_str)
    time_parts = time_str[:-2].split(":")
    hours = int(time_parts[0])
    minutes = int(time_parts[1])
    seconds = int(time_parts[2])
    am_pm = time_str[-2:]

    if hours == 12: hours = 0

    # Adjust hours if it's PM
    if am_pm.lower() == "pm":
        hours += 12

    # Calculate the total number of minutes
    total_minutes = (hours * 60) + minutes + (seconds // 60)
    
    return total_minutes

def save_to_single_file(workbook, current_date):
    # save to single files and remove from main file
    input_format = "%m/%d/%Y"
    output_format = "%d %b %Y"

    filename = "record.xlsx"

    date_obj = datetime.datetime.strptime(current_date, input_format)

    date_names = []

    for i in range(7):
        after_days = date_obj + datetime.timedelta(days = i)
        current_date = after_days.strftime(output_format)
        date_names.append(current_date)
        
    # workbook = load_workbook(filename)
    sheetnames = workbook.sheetnames

    print("SheetNames")
    print(sheetnames)
    print("DateNames")
    print(date_names)

    save_flag = True

    for sheetname in sheetnames:
        if sheetname in date_names:
            continue
        else:
            save_flag = False
            shutil.copy(filename, "temp.xlsx")
            workbook1 = load_workbook("temp.xlsx")

            workbook.remove_sheet(workbook.get_sheet_by_name(sheetname))

            sheets = workbook.sheetnames # ['Sheet1', 'Sheet2']

            for s in sheets:
                if s != sheetname:
                    sheet_name = workbook1.get_sheet_by_name(s)
                    workbook1.remove_sheet(sheet_name)
            save_workbook(sheetname.replace(" ", "_") + ".xlsx", workbook1)
            os.remove("temp.xlsx")

    # if save_flag == False: self.save_workbook("record.xlsx", workbook)

    print("SheetNames")
    print(workbook.sheetnames)



def ScrapeData():
    
    start_time = time.time()
    
    save_flag = False
    
    url = "https://www.m8huaythai.net"
    early_url1 = "https://m8huaythai.net/_View/RMOdds2.aspx?ot=e&ov=1&mt=0&wd="
    # early_url1 = "https://m8huaythai.net/_View/RMOdds2.aspx?update=false&r=316466324&wd="
    # "ot=e&ov=1&isWC=0&ia=0&isSiteFav=False"
    early_url2 = "&isWC=0&ia=0&tf=-1"
    today_url = "https://m8huaythai.net/_View/RMOdds2.aspx?ot=t&ov=1&mt=0&wd=&isWC=0&ia=0&tf=-1&isSiteFav=False"
    # "https://m8huaythai.net/_View/RMOdds2.aspx?update=false&r=316466324&wd=2023-10-31&ot=e&isWC=0&ia=0&isSiteFav=False"

    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_argument('--headless')
    chromeOptions.add_argument('--allow-running-insecure-content')
    chromeOptions.add_argument('--ignore-certificate-errors')
    
    chromeOptions.add_argument("enable-automation")
    chromeOptions.add_argument("--no-sandbox")
    chromeOptions.add_argument("--disable-extensions")
    chromeOptions.add_argument("--dns-prefetch-disable")
    chromeOptions.add_argument("--disable-gpu")

    # instantiate a webdriver
    driver = webdriver.Chrome(options = chromeOptions)

    # get date
    driver.get("https://www.m8huaythai.net/ClockTime.aspx")
    time.sleep(3)
    driver.execute_script("return document.readyState")
    time.sleep(2)

    date_html = driver.page_source
    # print(date_html)
    current_date_time = get_date_value(date_html)
    passed_minutes = calculate_passed_minutes(current_date_time[-10:])
    current_date = current_date_time[:10]

    print(current_date)
    print(passed_minutes)
    
    # First Login

    # if not os.path.exists("cookies.pkl"):
    driver.get(url)
    time.sleep(3)
    
    try:

        username_filed = driver.find_element('id', 'txtUserName')
        password_filed = driver.find_element('id', 'txtPassword')

        username_filed.send_keys(UserName)
        password_filed.send_keys(PassWord)
        
        password_filed.send_keys(Keys.RETURN)
    
    except NoSuchElementException:
        
        driver.close() # closing the webdriver 
        print("The website is not available now!")
        print("Please try again later...")
        
        end_time = time.time()

        return (end_time - start_time, passed_minutes)

    time.sleep(2)

    try:
        # After logging in
        cookies = driver.get_cookies()
    except Exception as e:
        print(e)
        print("The website is not available now!")
        print("Please try again later...")
        driver.close() # closing the webdriver 
        end_time = time.time()

        return (end_time - start_time, passed_minutes)


    # Store the cookies
    cookie_dict = {}
    for cookie in cookies:
        cookie_dict[cookie['name']] = cookie['value']

    time.sleep(5)

    while url == driver.current_url:
        time.sleep(1)

    print(driver.current_url)

    current_url = driver.current_url

    current_url = re.sub(r"lang=[A-Z]{2}-[A-Z]{2}", "lang=EN-US", current_url)
    
    driver.get(current_url)
    time.sleep(2)

    # Today Recording

    try:
        driver.get(today_url)

        # Add the stored cookies to the WebDriver
        for name, value in cookie_dict.items():
            driver.add_cookie({'name': name, 'value': value})

        driver.refresh()

    except Exception as e:
        print("Error: " + str(e))
    
    
    # this is just to ensure that the page is loaded 
    time.sleep(5) 
    
    driver.execute_script("return document.readyState")

    
    time.sleep(2)

    html = driver.page_source 
    # print(html)

    filename = "record.xlsx"

    analyzer = Analyzer(html, True, current_date, passed_minutes)
    leagues = analyzer.get_data()
    
    if not os.path.exists(filename):
        workbook = Workbook()
        save_workbook(filename, workbook)

    time_a = time.time()
    workbook = load_workbook(filename)
    time_b = time.time()
    print("Loading Excel costs " + str(time_b - time_a) + " seconds")

    save_to_single_file(workbook, current_date)

    if len(leagues) != 0:
        record = Record(current_date, passed_minutes)
        record.set_data(leagues)
        record.create_file_sheet(filename, workbook)

    # Early Recording

    k = 0
    
    for i in range(6):
        early_date_format = change_date_format(current_date, i, "%m/%d/%Y", "%Y-%m-%d")
        early_url = early_url1 + early_date_format + early_url2

        try:
            driver.get(early_url)

            # Add the stored cookies to the WebDriver
            for name, value in cookie_dict.items():
                driver.add_cookie({'name': name, 'value': value})

            driver.refresh()

        except Exception as e:
            print("Error: " + str(e))

        # this is just to ensure that the page is loaded 
        time.sleep(10) 

        html = driver.page_source 
        # print(html)
        
        driver.execute_script("return document.readyState")

        time.sleep(2)

        analyzer = Analyzer(html, False, current_date, passed_minutes)
        leagues = analyzer.get_data()

        if len(leagues) != 0:
            record = Record(current_date, passed_minutes)
            record.set_data(leagues)
            record.create_file_sheet(filename, workbook)
            k = k + 1
        if k == 5: break


    save_workbook(filename, workbook)
    
    print("URL")
    print(driver.current_url)

    # result match
    # check if there already exists result file.
    if passed_minutes > 120:
        previous_date = change_date_format(current_date, -1, "%m/%d/%Y", "%d %b %Y")
        recordedFileName = previous_date.replace(" ", "_")

        print("Previous_date " + recordedFileName + ".xlsx")

        if not os.path.exists(recordedFileName + "_result.xlsx"):
            if os.path.exists(recordedFileName + ".xlsx"):
                is_bug = False
                try:
                    workbook = load_workbook(recordedFileName + ".xlsx")
                    # workbook = load_workbook("record" + ".xlsx")
                    print(workbook.sheetnames)
                    worksheet = workbook[previous_date]
                    rows = (int(worksheet.max_row) - 3) // 24
                    print("Rows " + str(rows))
                    # if not save_flag:

                    row_base = 4
                    col_base = 7 # 6 -> 7 after inserting

                    worksheet.insert_cols(5, 1)
                    worksheet.column_dimensions['E'].width = 8
                    # merge cell
                    start_cell = 'A1'
                    end_cell = 'F2'

                    border_style = "thin"
                    border_color = "000000"

                    border = Border(left=Side(style=border_style, color = border_color),
                                    right=Side(style=border_style, color=border_color),
                                    top=Side(style=border_style, color=border_color),
                                    bottom=Side(style=border_style, color=border_color))
                    alignment = Alignment(horizontal = 'center', vertical = 'center')
                    font2 = Font(bold = True, name = 'Verdana', size = 7)
                    font1 = Font(bold = False, name = 'Verdana', size =7)

                    merge_range = f"{start_cell}:{end_cell}"
                    worksheet.merge_cells(merge_range)
                    
                    worksheet.cell(row = 3, column = 5).value = "Result"
                    worksheet.cell(row = 3, column = 5).font = font1

                    for i in range(rows * 24):
                        cell = worksheet.cell(row = row_base + i, column = 5)
                        cell.border = border
                        cell.font = font2
                        cell.alignment = alignment
                        if (i % 24) < 12: continue
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        
                    date_str = change_date_format(current_date, -1, "%m/%d/%Y", "%b%d%Y")
                    results = recordResult(date_str, passed_minutes)
                    result_map = {}
                    for result in results:
                        match_name = "".join(result.team1.split()) + "vs" + "".join(result.team2.split())
                        result_map[match_name.lower()] = str(result.score1) + "-" + str(result.score2)
                    print("ok1")
                    for i in range(rows):
                        idx = row_base + i * 24
                        cell = worksheet.cell(column = 2, row = idx)
                        cell_str = "".join(str(cell.value).split())
                        if cell_str == "": break
                        if cell_str.lower() in result_map:
                            pass
                        else: continue
                        score_str = result_map[cell_str.lower()]
                        # print("Score " + score_str)

                        score1 = float(score_str.split('-')[0])
                        score2 = float(score_str.split('-')[1])

                        isRefunded = 0
                        if score1 == 10000: isRefunded = 1

                        if isRefunded:
                            worksheet.cell(row = idx, column = 5).value = "Refunded"
                            worksheet.cell(row = idx + 12, column = 5).value = "Refunded"
                            continue
                        else:
                            worksheet.cell(row = idx, column = 5).value = score_str
                            worksheet.cell(row = idx + 12, column = 5).value = score_str

                        for ii in range(2):
                            start_range = col_base
                            end_range = int(worksheet.max_column)
                            start_col = -1
                            end_col = -1
                            for j in range(4):
                                idx = row_base + i * 24 + ii * 12 + j * 3
                                (start_col, end_col) = getRange(worksheet, idx, start_range, end_range)
                                # print("Start Col " + str(start_col))
                                # print("End Col " + str(end_col))
                                if start_col == -1 or end_col == -1: break
                                start_range = start_col
                                end_range = end_col
                                
                                markResultRow(worksheet, idx, score1, score2, start_col, end_col, ii)
                except Exception as e:
                    print(e)
                    is_bug = True
                    
                if not is_bug:
                    save_workbook(recordedFileName + "_result.xlsx", workbook)
                    # os.remove(recordedFileName + ".xlsx")

    driver.close() # closing the webdriver 

    end_time = time.time()

    return (end_time - start_time, passed_minutes)


# ScrapeData()

def markResultRow(worksheet, idx, score1, score2, start_col, end_col, ii):
    for i in range (start_col, end_col + 1):
        cell = worksheet.cell(row = idx + 1, column = i)
        cell_str = "".join(str(cell.value).split())
        # print("Cell Str " + cell_str)
        # print("Splitted " + cell_str.split("/")[0])
        if cell_str == "": continue
        if cell_str == "None": continue

        # print(cell_str)
        bet_result = float(cell_str.split("/")[0])
        if len(cell_str.split("/")) == 2:
            if bet_result < 0: bet_result = bet_result - 0.25
            elif bet_result > 0: bet_result = bet_result + 0.25
            else:
                if cell_str[0] == '-': bet_result = -0.25
                else: bet_result = 0.25

        res = 1
        if ii:
            if score1 + score2 > bet_result: res = 0
            elif score1 + score2 < bet_result: res = 2
        else:
            if score1 > score2 + bet_result: res = 0
            elif score1 < score2 + bet_result: res = 2
        # if res == 1: continue
            
        cell = worksheet.cell(row = idx + res, column = i)
        cell.fill = PatternFill(start_color="FECF81", end_color="FECF81", fill_type="solid")

def getRange(worksheet, idx, start_range, end_range):
    start_col = -1
    end_col = -1
    step = 1
    for j in range(start_range, end_range + 1, step):
        idy = j
        cell = worksheet.cell(row = idx, column = idy)
        if "".join(str(cell.value).split()) == "None" or "".join(str(cell.value).split()) == "": continue
        for k in range(step - 1, -1, -1):
            idk = idy - k
            if idk < start_range: continue
            cell = worksheet.cell(row = idx, column = idk)
            if "".join(str(cell.value).split()) == "None" or "".join(str(cell.value).split()) == "": continue
            start_col = idk
            break
        if start_col != -1: break
    for j in range(end_range, start_range - 1, -step):
        idy = j
        cell = worksheet.cell(row = idx, column = idy)
        if "".join(str(cell.value).split()) == "None" or "".join(str(cell.value).split()) == "": continue
        for k in range(step - 1, -1, -1):
            idk = idy + k
            if idk > end_range: continue
            cell = worksheet.cell(row = idx, column = idk)
            if "".join(str(cell.value).split()) == "None" or "".join(str(cell.value).split()) == "": continue
            end_col = idk
            break
        if end_col != -1: break
    return (start_col, end_col)


def recordResult(date_str, passed_minutes):
    url = "https://m8huaythai.net/_View/Result.aspx"

    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_argument('--headless')
    chromeOptions.add_argument('--allow-running-insecure-content')
    chromeOptions.add_argument('--ignore-certificate-errors')

    chromeOptions.add_argument("enable-automation")
    chromeOptions.add_argument("--no-sandbox")
    chromeOptions.add_argument("--disable-extensions")
    chromeOptions.add_argument("--dns-prefetch-disable")
    chromeOptions.add_argument("--disable-gpu")

    # instantiate a webdriver
    driver = webdriver.Chrome(options = chromeOptions)

    driver.get(url)

    results = []

    for i in range (2):
        if i == 1:
            lstDates = driver.find_element(By.NAME, "lstDates")
            lstDates.click()
            lstDates.send_keys(Keys.DOWN, Keys.ENTER)

        if passed_minutes > 700:
            lstDates = driver.find_element(By.NAME, "lstDates")
            lstDates.click()
            lstDates.send_keys(Keys.DOWN, Keys.ENTER)

        lstSortBy = driver.find_element(By.NAME, "lstSortBy")
        lstSortBy.click()

        lstSortBy.send_keys(Keys.DOWN, Keys.ENTER)

        driver.execute_script("return document.readyState")

        html = driver.page_source

        soup = BeautifulSoup(html, 'html.parser')

        table = soup.find('table', {'id': 'g1'})
        table_body = table.tbody

        tr_tag = table_body.tr

        isValidLeague = True

        while tr_tag.next_sibling is not None:
            if "".join(str(tr_tag.next_sibling.text).split()) == "": break
            tr_tag = tr_tag.next_sibling
            if 'style' in tr_tag.attrs:
                if "SABA" in tr_tag.td.span.text:
                    isValidLeague = False
                    continue
                elif "CORNER" in tr_tag.td.span.text:
                    isValidLeague = False
                    continue
                elif "OFFSIDE" in tr_tag.td.span.text:
                    isValidLeague = False
                    continue
                elif "BOOKING" in tr_tag.td.span.text:
                    isValidLeague = False
                    continue
                elif "TOTAL" in tr_tag.td.span.text:
                    isValidLeague = False
                    continue
                elif " - " in tr_tag.td.span.text:
                    isValidLeague = False
                    continue
                else:
                    isValidLeague = True
                    continue
            else:
                if isValidLeague == False: continue
                if "".join(str(tr_tag.td.next_sibling.next_sibling.next_sibling.next_sibling.text).replace('\n', '')) != "Completed": continue
                
                # # if date is valid
                # if date_str in "".join(str(tr_tag.td.text).replace('\n', '').split()) : pass
                # else: continue
                
                print("".join(str(tr_tag.td.next_sibling.next_sibling.next_sibling.next_sibling.text).replace('\n', '')) + ":")
                string = tr_tag.td.next_sibling.text
                string = " ".join(string.split())
                string_list = string.split(" -vs- ")
                first_team = string_list[0]
                second_team = string_list[1]
                print(first_team + ": VS :" + second_team)
                string = tr_tag.td.next_sibling.next_sibling.next_sibling.text
                isRefunded = 0
                if "Refund" in string: isRefunded = 1
                elif '-' not in string: continue
                first_value = 10000
                second_value = 10000
                if not isRefunded: 
                    string = " ".join(string.split())
                    string_list = string.split("-")
                    first_value = string_list[0]
                    second_value = string_list[1]
                print(str(first_value) + ": VS :" + str(second_value))
                result = MatchResult()
                result.team1 = first_team
                result.team2 = second_team
                result.score1 = first_value
                result.score2 = second_value
                results.append(result)
    driver.close()
    return results


def call_function(scheduler):

    runtime, record_time = ScrapeData()
    print("RECORD TIME: " + str(record_time))
    print("Runtime: " + str(runtime))
    # Calculate
    current_seconds = record_time * 60 + int(runtime) + 1 + 300
    next_time = current_seconds // 1800 + 1
    after = 1800 * next_time - current_seconds
    if after < 0: after = 0
    print("Record after " + str(after) + " seconds again")
    scheduler.enter(after, 1, call_function, (scheduler,))

my_scheduler = sched.scheduler(time.time, time.sleep)
my_scheduler.enter(1, 1, call_function, (my_scheduler,))
my_scheduler.run()
