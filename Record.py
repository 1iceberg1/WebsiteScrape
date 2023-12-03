import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import exceptions
from openpyxl.utils import cell as celll
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from datetime import datetime, timedelta
from openpyxl.formatting.rule import Rule

# import xlsxwriter

class Record():
    def __init__(self, current_date, passed_minutes):
        self.data = None
        self.current_date = current_date
        self.passed_minutes = passed_minutes
        # self.current_time = datetime.now().time()

        input_format = "%m/%d/%Y"
        output_format = "%d %b %Y"

        date_obj = datetime.strptime(self.current_date, input_format)
        today = date_obj.strftime(output_format)
        self.worksheet_name = today
        self.match_map = {}

    def check_timeline(self, worksheet):
        base = 6
        max_idx = (int(worksheet.max_column) - 5) // 48
        if max_idx < 0 : max_idx = 0
        for i in range(max_idx):
            date_value = str(worksheet.cell(row = 1, column = base + i * 48).value)
            if date_value == self.current_date:
                return (i, 1)
        return (max_idx, 0)
    
    def set_match_map(self, worksheet):
        self.match_map = {}
        base = 4
        # cnt = 0
        rows = (int(worksheet.max_row) - 3) // 24
        if rows < 0: rows = 0
        i = 0
        for i in range(rows):
            idx = base + i * 24
            cell = worksheet.cell(column = 2, row = idx)
            # print(str(cell.value))
            if "".join(str(cell.value).split()) == "": break
            self.match_map["".join(str(cell.value).split())] = i
    
    def set_data(self, leagues):
        self.data = leagues

    def save_workbook(self, filename, workbook):
        success_flag = True
        try:
            workbook.save(filename)
        except PermissionError:
            print("Failed to save the workbook. The file is currently open.")
            success_flag = False
        except exceptions.ReadOnlyWorkbookException:
            print("Failed to save the workbook. The file is opened as read-only.")
            success_flag = False
        except Exception as e:
            print("An error occurred while saving the workbook:", str(e))
            success_flag = False

        if success_flag == False:
            print("Please check your file is currently open or opend as read-only")

    def get_match_count(self, worksheet):
        num_rows = worksheet.max_row
        return (num_rows - 3) // 24
    
    def get_match_index(self, worksheet, match, league_name):
        base = 4
        # cnt = 0
        rows = (int(worksheet.max_row) - 3) // 24
        if rows < 0: rows = 0
        i = 0
        if "".join(str(match.match_name).split()) in self.match_map:
            i = self.match_map["".join(str(match.match_name).split())]
            idx = base + i * 24
            if "".join(str(worksheet.cell(row = idx, column = 1).value).split()) == "".join(str(league_name).split()):
                if "".join(str(worksheet.cell(row = idx, column = 4).value).split()) == "".join(str(match.time).split()):
                    return (i, 1)
                else:
                    for j in range(24):
                        worksheet.cell(row = idx + j, column = 4).value = str(match.time)
                    ret = self.replace_match(worksheet, match, i)
                    return (ret, 0)
        ret, f = self.insert_match(worksheet, match)
        return (ret, f)

    def check_match_time(self, prev_match_time, match_time):

        time_format = "%I:%M%p"

        try:
            # Convert the time strings to datetime objects
            time1 = datetime.strptime(prev_match_time, time_format)
            time2 = datetime.strptime(match_time, time_format)

            # Cacluate the time difference in minutes
            time_diff = (time2 - time1).total_seconds() / 60
            
        except:
            print(prev_match_time)
            print(match_time)
            print(self.current_date)
        return time_diff

    def insert_match(self, worksheet, match):
        rows = (int(worksheet.max_row) - 3) // 24
        if rows < 0: rows = 0
        i = 0
        base = 4
        st = 0
        en = rows - 1
        mid = 0
        flag = 0
        while st <= en:
            mid = (st + en) // 2
            idx = base + mid * 24
            prev_time = str(worksheet.cell(column = 4, row = idx).value)
            time_diff = self.check_match_time(prev_time, match.time)
            
            if time_diff == 0:
                flag = 1
                break
            if time_diff > 0:
                st = mid + 1
            if time_diff < 0:
                en = mid - 1

        if flag == 1: i = mid
        else: i = st
        idx = base + i * 24
        worksheet.insert_rows(idx, 24)
        if i == rows: 
            return (i, 2)
        else: return (i, 0)

    def replace_match(self, worksheet, match, n):
        rows = (int(worksheet.max_row) - 3) // 24
        cols = int(self.get_max_column(worksheet))
        if rows < 0: rows = 0
        i = 0
        base = 4

        for i in range(rows):
            if i == n: continue
            idx = base + i * 24
            prev_time = str(worksheet.cell(column = 4, row = idx).value)
            print("Prev: " + prev_time + " " + str(idx))
            if self.check_match_time(prev_time, match.time) < 0: break
        else: 
            if rows: i = i + 1
        idx = base + i * 24
        worksheet.insert_rows(idx, 24)
        pivot_n = n
        
        if i < n: pivot_n = pivot_n + 1
        pivot_idx = base + pivot_n * 24
        col_letter = celll.get_column_letter(cols)
        range_string = "A" + str(pivot_idx) + ":" + col_letter + str(pivot_idx + 23)

        worksheet.move_range(range_string, rows = (i - pivot_n) * 24)
        worksheet.delete_rows(pivot_idx, 24)
        return i
    
    def get_max_column(self, worksheet):
        cols = (int(worksheet.max_column) - 5) // 48
        i = 0
        for i in range(cols):
            if "".join(str(worksheet.cell(column = 6 + i * 48, row = 2).value)) == "":
                break
        else: 
            if cols: i = i + 1
        return i * 48 + 5

    def get_max_row(self, worksheet):
        rows = (int(worksheet.max_row) - 3) // 24
        i = 0
        for i in range(rows):
            if "".join(str(worksheet.cell(column = 4, row = 4 + i * 24).value)) == "":
                break
        else: 
            if rows: i = i + 1
        return i * 24 + 3
    
    def check_match(self, worksheet, match, league_name):
        base = 4
        # cnt = 0
        rows = worksheet.max_row
        for i in range((rows - 3) // 24):
            idx = base + i * 24
            cell = worksheet.cell(column = 2, row = idx)
            # print(str(cell.value))
            if "".join(str(cell.value).split()) == "": break
            if "".join(str(cell.value).split()) == "".join(str(match.match_name).split()):
                if "".join(str(worksheet.cell(column = 1, row = idx).value).split()) == "".join(str(league_name).split()):
                    # print("Checked: " + str(idx))
                    return i
            # cnt = cnt + 1
        # print("Checked: -1")
        return -1
            
    def create_match(self, worksheet, i, day_diff, match, league_name):
        
        base = 4 + i * 24
        font = Font(bold = False, name = 'Verdana', size = 7)
        font1 = Font(bold = False, name = 'Verdana', size = 8)
        font2 = Font(bold = True, name = 'Verdana', size = 8)
        alignment = Alignment(horizontal = 'center', vertical = 'center')
        pattern = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        border_style = "thin"
        border_color = "000000"

        border = Border(left=Side(style=border_style, color = border_color),
                        right=Side(style=border_style, color=border_color),
                        top=Side(style=border_style, color=border_color),
                        bottom=Side(style=border_style, color=border_color))

        passed = self.passed_minutes - 25
        if passed < 0: passed = 0
        index = passed // 30 + day_diff * 48

        str_check = str(worksheet.cell(row = base, column = 1).value)
        check_cell = worksheet.cell(row = base, column = 6 + day_diff * 48)

        # print("CHECK: " + str_check)

        for j in range(24):
            if "".join(str_check.split()) == "" or str_check == "None":
                if j == 0: print("NEW!")
            
                worksheet.cell(row = base + j, column = 1).value = league_name
                worksheet.cell(row = base + j, column = 2).value = match.match_name
                worksheet.cell(row = base + j, column = 3).value = match.date
                worksheet.cell(row = base + j, column = 4).value = match.time

                cell = worksheet.cell(row = base + j, column = 5)
                cell.font = font1
                
                if j % 3 == 0:
                    if j < 12:
                        cell.value = "H"
                    else:
                        cell.value = "O"
                elif j % 3 == 1:
                    if (j % 12) // 3 == 0:
                        cell.value = "MP"
                    else:
                        cell.value = str((j % 12) // 3) + "P"
                    cell.font = font2
                else:
                    if j < 12:
                        cell.value = "A"
                    else:
                        cell.value = "U"

                for k in range(5):
                    cell1 = worksheet.cell(row = base + j, column = k + 1)
                    cell1.font = font
                    if k == 4:
                        if j % 3 == 1: cell1.font = font2
                        else: cell1.font = font1
                    cell1.alignment = alignment
                    cell1.border = border
                    if j < 12: continue
                    cell1.fill = pattern
            
            if check_cell.fill != pattern:
                for k in range(48 * day_diff + 5, 48 * day_diff + 53):
                    cell1 = worksheet.cell(row = base + j, column = k + 1)
                    if j % 3 == 1: cell1.font = font2
                    else: cell1.font = font1
                    cell1.alignment = alignment
                    cell1.border = border
                    if j < 12: continue
                    cell1.fill = pattern
        
        worksheet.cell(row = base + 0, column = 6 + index).value = match.h1
        worksheet.cell(row = base + 1, column = 6 + index).value = match.hp1
        worksheet.cell(row = base + 2, column = 6 + index).value = match.a1
        worksheet.cell(row = base + 3, column = 6 + index).value = match.h2
        worksheet.cell(row = base + 4, column = 6 + index).value = match.hp2
        worksheet.cell(row = base + 5, column = 6 + index).value = match.a2
        worksheet.cell(row = base + 6, column = 6 + index).value = match.h3
        worksheet.cell(row = base + 7, column = 6 + index).value = match.hp3
        worksheet.cell(row = base + 8, column = 6 + index).value = match.a3
        worksheet.cell(row = base + 9, column = 6 + index).value = match.h4
        worksheet.cell(row = base + 10, column = 6 + index).value = match.hp4
        worksheet.cell(row = base + 11, column = 6 + index).value = match.a4

        worksheet.cell(row = base + 12, column = 6 + index).value = match.o1
        worksheet.cell(row = base + 13, column = 6 + index).value = match.op1
        worksheet.cell(row = base + 14, column = 6 + index).value = match.u1
        worksheet.cell(row = base + 15, column = 6 + index).value = match.o2
        worksheet.cell(row = base + 16, column = 6 + index).value = match.op2
        worksheet.cell(row = base + 17, column = 6 + index).value = match.u2
        worksheet.cell(row = base + 18, column = 6 + index).value = match.o3
        worksheet.cell(row = base + 19, column = 6 + index).value = match.op3
        worksheet.cell(row = base + 20, column = 6 + index).value = match.u3
        worksheet.cell(row = base + 21, column = 6 + index).value = match.o4
        worksheet.cell(row = base + 22, column = 6 + index).value = match.op4
        worksheet.cell(row = base + 23, column = 6 + index).value = match.u4

    def formatting_sheet_first_time(self, worksheet):
        # Formatting Worksheet For the First Time

        # merge sell
        start_cell = 'A1'
        end_cell = 'E2'

        merge_range = f"{start_cell}:{end_cell}"
        worksheet.merge_cells(merge_range)

        # set the value of merge cell
        merge_cell = worksheet[start_cell]
        merge_cell.value = "Timestamp of recording"

        # set the background color of the merge cell
        merge_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # set the font style, size, and alignment
        merge_font = Font(bold=False, name='Verdana', size=7)
        merge_alignment = Alignment(horizontal='center', vertical='center')
        merge_cell.font = merge_font
        merge_cell.alignment = merge_alignment

        # format title cells
        worksheet.cell(row = 3, column = 1).value = 'League'
        worksheet.cell(row = 3, column = 1).font = Font(bold=False, name='Verdana', size=7)
        worksheet.cell(row = 3, column = 1).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.cell(row = 3, column = 2).value = 'Match'
        worksheet.cell(row = 3, column = 2).font = Font(bold=False, name='Verdana', size=7)
        worksheet.cell(row = 3, column = 2).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.cell(row = 3, column = 3).value = 'Start Date'
        worksheet.cell(row = 3, column = 3).font = Font(bold=False, name='Verdana', size=7)
        worksheet.cell(row = 3, column = 3).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.cell(row = 3, column = 4).value = 'Start Time'
        worksheet.cell(row = 3, column = 4).font = Font(bold=False, name='Verdana', size=7)
        worksheet.cell(row = 3, column = 4).alignment = Alignment(horizontal='center', vertical='center')

        # set column width
        worksheet.column_dimensions['A'].width = 39
        worksheet.column_dimensions['B'].width = 39
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 8
        worksheet.column_dimensions['E'].width = 3

    def calculate_days_between_dates(self, date_str1, date_str2):
        # Define the input date format
        input_format = "%m/%d/%Y"

        # Convert the date strings to datetime objects
        date_obj1 = datetime.strptime(date_str1, input_format)
        date_obj2 = datetime.strptime(date_str2, input_format)

        # Calculate the difference between the two dates
        diff = date_obj2 - date_obj1

        # Extract the number of days from the difference
        days = diff.days

        return days
    
    def adding_timeline(self, worksheet, idx):
                
        # formatting timeline
        for i in range(48 * idx, 48 * idx + 48):
            worksheet.cell(row = 1, column = 6 + i).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            worksheet.cell(row = 2, column = 6 + i).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            worksheet.column_dimensions[celll.get_column_letter(6 + i)].width = 7.5

        font = Font(bold=False, name='Verdana', size=7)

        for i in range(48):
            hour = i // 2
            if hour >= 12: hour = hour - 12
            if hour == 0: hour = 12
            minute = (i % 2) * 30 + 25
            ap = "AM"
            if i >= 24:
                ap = "PM"
            value = str(hour) + ":" + str(minute).zfill(2) + " " + ap
            
            cell = worksheet.cell(row = 2, column = 48 * idx + i + 6)
            cell.value = value
            cell.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
            cell.font = font
            cell = worksheet.cell(row = 1, column = 48 * idx + i + 6)
            cell.value = self.current_date
            cell.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
            cell.font = font

    def create_file_sheet(self, filename):

        create_flag = False

        if not os.path.exists(filename):
            workbook = Workbook()
            self.save_workbook(filename, workbook)
            create_flag = True

        workbook = load_workbook(filename)

        # Generate the name with today's date

        input_format = "%m/%d/%Y"
        output_format = "%d %b %Y"

        date_obj = datetime.strptime(self.current_date, input_format)
        today = date_obj.strftime(output_format)
        save_flag = False

        print(today)

        date_names = []

        for i in range(7):
            after_days = date_obj + timedelta(days = i)

            current_date = after_days.strftime(output_format)
            date_names.append(current_date)
            isNew = False

            if current_date in workbook.sheetnames:
                print('Exists')
            else:
                print("No")
                workbook.create_sheet(title = current_date)
                isNew = True
                save_flag = True

            worksheet = workbook[current_date]
            if isNew:
                self.formatting_sheet_first_time(worksheet)

        if save_flag:
            self.save_workbook(filename, workbook)

        # save to single files and remove from main file
        
        workbook = load_workbook(filename)

        for sheetname in workbook.sheetnames:
            if sheetname in date_names:
                continue
            else:
                sheets = workbook.sheetnames # ['Sheet1', 'Sheet2']

                for s in sheets:
                    if s != sheetname:
                        sheet_name = workbook.get_sheet_by_name(s)
                        workbook.remove_sheet(sheet_name)
                self.save_workbook(sheetname + ".xlsx", workbook)

        workbook = load_workbook(filename)

        # get data for processing
        leagues = self.data
        
        # update match record

        # initialize updating variables
        compare_date = leagues[0].matches[0].date
        update_flag = False
        cnt = 0
        date_idx = 0

        worksheet  = None

        for league in leagues:
            matches = league.matches
            for match in matches:
                
                match_date = match.date
                
                if match_date != compare_date:
                    compare_date = match_date
                    update_flag = False
                    cnt = 0
                
                if update_flag == False:
                    date_obj = datetime.strptime(match_date, "%m/%d/%Y")
                    worksheet = workbook[date_obj.strftime("%d %b %Y")]

                    date_idx, isExist = self.check_timeline(worksheet)
                    
                    if isExist == False:
                        self.adding_timeline(worksheet, date_idx)

                    self.set_match_map(worksheet)

                    # get amount of current recoreded matches
                    cnt = self.get_match_count(worksheet)
                    print("current matches " + str(cnt))

                    update_flag = True

                    # idx = self.check_match(worksheet, match, league.league_name)
                    # update_flag = True
                    # if idx != -1:
                    #     cnt = idx

                idx, f = self.get_match_index(worksheet, match, league.league_name)
                
                # self.create_match(worksheet, cnt, date_idx, match, league.league_name)
                self.create_match(worksheet, idx, date_idx, match, league.league_name)
                if f == 0: self.set_match_map(worksheet)
                if f == 2:
                    rows = (int(worksheet.max_row) - 3) // 24
                    self.match_map["".join(str(match.match_name).split())] = rows - 1
                cnt = cnt + 1

        self.save_workbook(filename, workbook)
